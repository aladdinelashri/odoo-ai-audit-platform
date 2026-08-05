"""
Export generators for reports: JSON, Excel (.xlsx), and PDF (.pdf).
Uses openpyxl (write-only for large datasets) and reportlab.
"""
import io
import json
from typing import List, Optional, Any, Dict
import sqlite3

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
import logging

logger = logging.getLogger(__name__)


def export_json(rows: List[sqlite3.Row], columns: Optional[List[str]] = None) -> str:
    """
    Export data to a JSON string (pretty-printed).
    If columns are provided, only those fields are included.
    """
    row_dicts = _normalize_rows(rows)
    if columns is None and row_dicts:
        columns = _get_columns(row_dicts)
    # If columns provided, filter each row to those keys
    if columns:
        filtered = [{col: row.get(col, None) for col in columns} for row in row_dicts]
    else:
        filtered = row_dicts
    return json.dumps(filtered, indent=2, default=str)


def _normalize_rows(rows: List[sqlite3.Row]) -> List[Dict[str, Any]]:
    """Convert sqlite3.Row objects to dictionaries if needed."""
    if rows and hasattr(rows[0], 'keys'):
        return [dict(row) for row in rows]
    return rows  # assume they are already dicts


def _get_columns(rows: List[sqlite3.Row]) -> List[str]:
    """Extract column names from the first row."""
    if not rows:
        return []
    if hasattr(rows[0], 'keys'):
        return list(rows[0].keys())
    # If rows are dicts, get keys
    if isinstance(rows[0], dict):
        return list(rows[0].keys())
    # Fallback: if rows are lists/tuples, use generic names
    return [f"col_{i}" for i in range(len(rows[0]))] if rows else []


def export_excel(rows: List[sqlite3.Row], columns: Optional[List[str]] = None, sheet_name: str = "Report") -> bytes:
    """
    Export data to an Excel .xlsx file using write-only mode for memory efficiency.
    Returns bytes of the file.
    """
    if not rows:
        # Return an empty workbook with a header row
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns if columns else ["No Data"])
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Convert to dict if needed
    row_dicts = _normalize_rows(rows)
    if columns is None:
        columns = _get_columns(rows)

    # Write in write-only mode
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)

    # Write header
    ws.append(columns)

    # Write data rows
    for row in row_dicts:
        # Ensure row values are in column order
        ws.append([row.get(col, None) for col in columns])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_pdf(rows: List[sqlite3.Row], columns: Optional[List[str]] = None, title: str = "Report Export") -> bytes:
    """Export data to PDF. Limits to 1000 rows to prevent server hang."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        title=title,
    )

    font_name = 'Helvetica'
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=0.5 * inch,
        fontName=font_name,
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.25 * inch))

    if not rows:
        elements.append(Paragraph("No data available.", styles['Normal']))
        doc.build(elements)
        return buffer.getvalue()

    row_dicts = _normalize_rows(rows)
    if columns is None:
        columns = _get_columns(rows)

    total_rows = len(row_dicts)
    max_rows = 1000  # CAP to prevent server hang
    displayed_rows = row_dicts[:max_rows]

    # Add summary
    elements.append(Paragraph(
        f"<b>Total rows:</b> {total_rows} | <b>Displayed:</b> {len(displayed_rows)}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.25 * inch))

    # Build table
    table_data = [columns]
    for row in displayed_rows:
        table_data.append([str(row.get(col, '')) for col in columns])

    num_cols = len(columns)
    available_width = doc.width
    col_widths = [available_width / num_cols for _ in columns]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    tbl_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name + '-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(tbl_style)
    elements.append(table)

    def _page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.drawRightString(doc.width + doc.leftMargin, doc.bottomMargin - 0.3 * inch,
                               f"Page {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_page_number, onLaterPages=_page_number)
    return buffer.getvalue()

    return buffer.getvalue()
